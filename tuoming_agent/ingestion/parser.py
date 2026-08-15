from __future__ import annotations

import codecs
from collections.abc import Iterator
from dataclasses import dataclass
from io import BytesIO
from itertools import islice
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd

MISSING_TEXT_MARKERS = frozenset({"nan", "n/a", "null", "none"})


class UnsupportedFileError(ValueError):
    """Raised for unsupported or unreadable uploads."""


@dataclass(frozen=True)
class ParsedTable:
    logical_name: str
    sheet_name: str | None
    dataframe: pd.DataFrame


def iter_file_chunks(
    filename: str, source: BinaryIO, chunk_rows: int = 50_000
) -> Iterator[ParsedTable]:
    if chunk_rows < 1:
        raise ValueError("chunk_rows must be positive.")
    suffix = Path(filename).suffix.lower()
    stem = Path(filename).stem.strip() or "dataset"
    if suffix == ".csv":
        encoding = _detect_csv_encoding(source)
        source.seek(0)
        try:
            numeric_candidates: set[str] | None = None
            columns_with_values: set[str] = set()
            for dataframe in pd.read_csv(
                source,
                encoding=encoding,
                chunksize=chunk_rows,
                skip_blank_lines=False,
                keep_default_na=False,
                dtype=str,
            ):
                normalized = _normalize_missing_values(dataframe)
                if numeric_candidates is None:
                    numeric_candidates = set(normalized.columns)
                for column in tuple(numeric_candidates):
                    present = normalized[column].notna()
                    if not present.any():
                        continue
                    columns_with_values.add(column)
                    numeric = pd.to_numeric(normalized.loc[present, column], errors="coerce")
                    if numeric.isna().any():
                        numeric_candidates.discard(column)

            numeric_columns = (numeric_candidates or set()) & columns_with_values
            source.seek(0)
            for dataframe in pd.read_csv(
                source,
                encoding=encoding,
                chunksize=chunk_rows,
                skip_blank_lines=False,
                keep_default_na=False,
                dtype=str,
            ):
                yield ParsedTable(
                    stem,
                    None,
                    _normalize_missing_values(
                        dataframe, numeric_columns=numeric_columns
                    ),
                )
        except (UnicodeDecodeError, pd.errors.ParserError) as exc:
            raise UnsupportedFileError("CSV structure could not be parsed.") from exc
        return
    if suffix in {".xlsx", ".xlsm"}:
        try:
            source.seek(0)
            workbook = openpyxl.load_workbook(
                source, read_only=True, data_only=True, keep_vba=False
            )
            try:
                for worksheet in workbook.worksheets:
                    rows = worksheet.iter_rows(values_only=True)
                    header = next(rows, None)
                    if header is None:
                        continue
                    columns = _normalize_excel_headers(header)
                    while batch := list(islice(rows, chunk_rows)):
                        dataframe = pd.DataFrame(batch, columns=columns)
                        yield ParsedTable(
                            f"{stem}::{worksheet.title}",
                            worksheet.title,
                            _normalize_missing_values(dataframe),
                        )
            finally:
                workbook.close()
        except Exception as exc:
            if isinstance(exc, UnsupportedFileError):
                raise
            raise UnsupportedFileError("Excel workbook could not be parsed.") from exc
        return
    raise UnsupportedFileError("Only CSV, XLSX and XLSM files are supported.")


def parse_file(filename: str, content: bytes) -> list[ParsedTable]:
    suffix = Path(filename).suffix.lower()
    stem = Path(filename).stem.strip() or "dataset"
    if suffix == ".csv":
        return [ParsedTable(stem, None, _read_csv(content))]
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(stem, content)
    raise UnsupportedFileError("Only CSV, XLSX and XLSM files are supported.")


def preview_file(
    filename: str, source: BinaryIO, sample_rows: int = 500
) -> list[ParsedTable]:
    """Read only the first rows of each uploaded table for policy selection."""
    suffix = Path(filename).suffix.lower()
    stem = Path(filename).stem.strip() or "dataset"
    if suffix == ".csv":
        return [ParsedTable(stem, None, _read_csv(source, sample_rows))]
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(stem, source, sample_rows)
    raise UnsupportedFileError("Only CSV, XLSX and XLSM files are supported.")


def _read_csv(content: bytes | BinaryIO, sample_rows: int | None = None) -> pd.DataFrame:
    errors: list[str] = []
    source = BytesIO(content) if isinstance(content, bytes) else content
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "big5"):
        try:
            source.seek(0)
            dataframe = pd.read_csv(
                source,
                encoding=encoding,
                nrows=sample_rows,
                skip_blank_lines=False,
                keep_default_na=False,
                dtype=str,
            )
            return _normalize_missing_values(dataframe, infer_numeric_strings=True)
        except UnicodeDecodeError:
            errors.append(encoding)
        except pd.errors.ParserError as exc:
            raise UnsupportedFileError("CSV structure could not be parsed.") from exc
    raise UnsupportedFileError(f"CSV encoding is unsupported; attempted: {', '.join(errors)}")


def _detect_csv_encoding(source: BinaryIO, sample_size: int = 64 * 1024) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "big5"):
        source.seek(0)
        decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
        try:
            while chunk := source.read(sample_size):
                decoder.decode(chunk, final=False)
            decoder.decode(b"", final=True)
            return encoding
        except UnicodeDecodeError:
            continue
    raise UnsupportedFileError("CSV encoding is unsupported.")


def _normalize_excel_headers(header: tuple[object, ...]) -> list[str]:
    bases = [
        f"Unnamed: {index}" if value is None or str(value) == "" else str(value)
        for index, value in enumerate(header)
    ]
    columns: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(header):
        base = bases[index]
        candidate = base
        if candidate in used or (value is None and candidate in bases[index + 1 :]):
            suffix = 1
            candidate = f"{base}.{suffix}"
            reserved = set(bases[index + 1 :])
            while candidate in used or candidate in reserved:
                suffix += 1
                candidate = f"{base}.{suffix}"
        columns.append(candidate)
        used.add(candidate)
    return columns


def _normalize_missing_values(
    dataframe: pd.DataFrame,
    *,
    infer_numeric_strings: bool = False,
    numeric_columns: set[str] | None = None,
) -> pd.DataFrame:
    def normalize(value: object) -> object:
        if not isinstance(value, str):
            return value
        stripped = value.strip()
        if not stripped or stripped.casefold() in MISSING_TEXT_MARKERS:
            return pd.NA
        return value

    normalized = dataframe.map(normalize).infer_objects()
    if not infer_numeric_strings and numeric_columns is None:
        return normalized
    columns = normalized.columns if numeric_columns is None else numeric_columns
    for column in columns:
        series = normalized[column]
        if not (
            pd.api.types.is_object_dtype(series.dtype)
            or pd.api.types.is_string_dtype(series.dtype)
        ):
            continue
        present = series.notna()
        if not present.any():
            continue
        numeric = pd.to_numeric(series, errors="coerce")
        if numeric[present].notna().all():
            normalized[column] = numeric
    return normalized.infer_objects()


def _read_excel(
    stem: str, content: bytes | BinaryIO, sample_rows: int | None = None
) -> list[ParsedTable]:
    try:
        source = BytesIO(content) if isinstance(content, bytes) else content
        source.seek(0)
        workbook = openpyxl.load_workbook(
            source, read_only=True, data_only=True, keep_vba=False
        )
        try:
            tables: list[ParsedTable] = []
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(values_only=True)
                header = next(rows, None)
                if header is None:
                    continue
                dataframe = pd.DataFrame(
                    list(islice(rows, sample_rows)), columns=_normalize_excel_headers(header)
                )
                dataframe = _normalize_missing_values(dataframe)
                tables.append(
                    ParsedTable(
                        logical_name=f"{stem}::{worksheet.title}",
                        sheet_name=worksheet.title,
                        dataframe=dataframe,
                    )
                )
            return tables
        finally:
            workbook.close()
    except Exception as exc:
        raise UnsupportedFileError("Excel workbook could not be parsed.") from exc
